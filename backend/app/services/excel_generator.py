"""Generate the BOM Excel file from a QuotationResponse."""
from io import BytesIO
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.models.schemas import QuotationResponse


HEADER_FILL = PatternFill("solid", start_color="1F4E78", end_color="1F4E78")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(name="Arial", bold=True, size=14)
SUBTITLE_FONT = Font(name="Arial", size=10, italic=True, color="595959")
NORMAL_FONT = Font(name="Arial", size=10)
BOLD_FONT = Font(name="Arial", size=10, bold=True)
WARNING_FONT = Font(name="Arial", size=9, italic=True, color="C00000")

THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

CURRENCY_FMT = '"$"#,##0.00;("$"#,##0.00);"-"'


def generate_bom_excel(quotation: QuotationResponse) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "BOM"

    # --- Title block ---
    ws["A1"] = "AC COMBINER PANEL - BILL OF MATERIALS"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:F1")

    ws["A2"] = f"Customer: {quotation.customer_name}"
    ws["A2"].font = SUBTITLE_FONT
    ws["A3"] = f"Project: {quotation.project_name}"
    ws["A3"].font = SUBTITLE_FONT
    ws["A4"] = f"OEM: {quotation.oem}"
    ws["A4"].font = SUBTITLE_FONT
    ws["A5"] = f"Date: {date.today().isoformat()}"
    ws["A5"].font = SUBTITLE_FONT

    header_row = 7
    headers = ["S/N", "Item Description", "Quantity", "Part Number", "Unit Price ($)", "Total Price ($)"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER

    row = header_row + 1
    for line in quotation.bom:
        ws.cell(row=row, column=1, value=line.sn).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=2, value=line.description)
        ws.cell(row=row, column=3, value=line.quantity).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=4, value=line.part_number).alignment = Alignment(horizontal="center")

        unit_cell = ws.cell(row=row, column=5, value=line.unit_price)
        unit_cell.number_format = CURRENCY_FMT

        total_cell = ws.cell(row=row, column=6, value=f"=C{row}*E{row}")
        total_cell.number_format = CURRENCY_FMT

        for col in range(1, 7):
            cell = ws.cell(row=row, column=col)
            cell.font = NORMAL_FONT
            cell.border = BORDER
            if col == 2:
                cell.alignment = Alignment(wrap_text=True, vertical="center")

        row += 1

    # --- Grand total row ---
    total_label_cell = ws.cell(row=row, column=5, value="GRAND TOTAL")
    total_label_cell.font = BOLD_FONT
    total_label_cell.alignment = Alignment(horizontal="right")
    total_label_cell.border = BORDER

    first_data_row = header_row + 1
    last_data_row = row - 1
    grand_total_cell = ws.cell(row=row, column=6, value=f"=SUM(F{first_data_row}:F{last_data_row})")
    grand_total_cell.font = BOLD_FONT
    grand_total_cell.number_format = CURRENCY_FMT
    grand_total_cell.border = BORDER
    grand_total_cell.fill = PatternFill("solid", start_color="FFF2CC", end_color="FFF2CC")

    for col in range(1, 5):
        ws.cell(row=row, column=col).border = BORDER

    row += 2

    # --- Warnings / notes ---
    if quotation.warnings:
        ws.cell(row=row, column=1, value="Notes:").font = BOLD_FONT
        row += 1
        for w in quotation.warnings:
            cell = ws.cell(row=row, column=1, value=f"• {w}")
            cell.font = WARNING_FONT
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
            cell.alignment = Alignment(wrap_text=True)
            row += 1

    # --- Column widths ---
    widths = {1: 6, 2: 55, 3: 10, 4: 20, 5: 14, 6: 14}
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.freeze_panes = f"A{header_row + 1}"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
